from csv import Sniffer, DictReader
from openpyxl import load_workbook as lw


class Reader:
    """
    Base Reader class
    """
    def check_template(self, snippet):
        """
        Checks if snippet needs to be personalized for each recipient.
        Changes self.required to True if personalization is required.

        Returns True if personalization needed else False
        """
        for header in self.headers:
            if f"{{{{ {header} }}}}" in snippet.content:
                self.required = True
                break
        else:
            return False

        return True
    
    def get_headers(self):
        """
        Returns self.headers
        """
        return self.headers

    def is_required(self):
        """
        Returns self.required
        """
        return self.required


class CSVReader(Reader):
    """
    This class inherits from mailtorpedo.reader.Reader object and processes 
    CSV/TSV or similar files. Returns a list of dict instances having column 
    headers as key and value for each row.
    """
    def __init__(self, filename, email_field, encoding="utf-8"):
        """
        filename: File path (instance of str class)
            Path of the file that contains delimeter separated values
        email_field: str
            Name of the column header that contains the email addresses of the recipient
        encoding (optional): str
            Encoding of the file mentioned as filename
        """
        self.filename = filename
        self.encoding = encoding
        self.required = False
        self.email_field = email_field

        with open(self.filename, "r", encoding=self.encoding) as csvfile:
            topline = csvfile.readline().rstrip()
            self.delimiter = Sniffer().sniff(topline).delimiter

        headers = topline.split(self.delimiter)
        self.headers = headers

    def parsed_dict(self):
        """
        Returns a list having dictionary associated for each row of delimiter seperated values.

        Each of the dictionaries represent a row in CSV file where the keys are column headers
        and the values are associated cells in each column for that row.
        """
        if self.email_field not in self.headers:
            raise ValueError(f"No column with header '{self.email_field}' is present")

        with open(self.filename, "r", encoding=self.encoding) as csvfile:
            reader = list(DictReader(csvfile, delimiter=self.delimiter))
        return reader


class ExcelReader(Reader):
    """
    This class inherits from mailtorpedo.reader.Reader class and processes 
    Microsoft Excel files. Returns a list of dict instances having column 
    headers as key and value for each row.
    """
    def __init__(self, filename, email_field, sheet=0):
        """
        filename: File path (instance of str class)
            Path of the file that contains delimeter separated values
        email_field: str
            Name of the column header that contains the email addresses of the recipient
        sheet (optional): int or str
            Name of the worksheet in excel file or its serial number (indexed from 0)
            By default it opens the first sheet of the excel file.
        """
        self.filename = filename
        self.email_field = email_field
        self.sheet_index = sheet
        self.required = False
        self.load_workbook()

    def load_workbook(self):
        """
        Loads the Excel file and processes it.
        
        Returns None
        """
        self.workbook = lw(self.filename, read_only=True, data_only=True)
        if type(self.sheet_index) == int:
            self.workbook.active = self.sheet_index
        elif type(self.sheet_index) == str:
            if self.sheet_index in self.workbook.sheetnames:
                self.workbook.active = self.workbook.sheetnames.index(self.sheet_index)
            else:
                raise Exception("Sheet not found")
        else:
            raise TypeError("Sheet value must be integer or string")
        self.sheet = self.workbook.active
        self.headers = [i.value for i in self.workbook.active[1]]

    def change_sheet(self, sheet):
        self.sheet_index = sheet
        if type(self.sheet_index) == int:
            self.workbook.active = self.sheet_index
        elif type(self.sheet_index) == str:
            if self.sheet_index in self.workbook.sheetnames:
                self.workbook.active = self.workbook.sheetnames.index(self.sheet_index)
            else:
                raise Exception("Sheet not found")
        else:
            raise TypeError("Sheet value must be integer or string")
        self.sheet = self.workbook.active
        self.headers = [i.value for i in self.workbook.active[1]]

    def parsed_dict(self):
        """
        Returns a list having dictionary associated for each row of delimiter seperated values.

        Each of the dictionaries represent a row in the Excel sheet where the keys are column 
        headers and the values are associated cells in each column for that row.
        """
        if self.email_field not in self.headers:
            raise ValueError(f"No column with header '{self.email_field}' is present")

        header_length = len(self.headers)
        reader = []
        rowtuples = self.sheet[2 : self.sheet.max_row]
        if type(rowtuples[0]).__name__ != "tuple":
            rowtuples = (rowtuples,)
        for i in rowtuples:
            temp_dict = dict()
            for k in range(0, header_length):
                if self.headers[k] == self.email_field and i[k].value is None:
                    break
                temp_dict[self.headers[k]] = str(i[k].value)
            else:
                reader.append(temp_dict)
        
        self.workbook.close()

        return reader
